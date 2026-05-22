from __future__ import annotations

from itertools import product
from pathlib import Path
import argparse

import pandas as pd
from openpyxl import load_workbook

import q1_solution

"""
q4_solution.py

用途：基于 q2/q1 输出，重新搜索最优统一 markup 定价方案。
通过顶部常量可以快速修改：MARKUP_OPTIONS、CAP_PER_SIZE、PROFIT_RATE_LIMIT。
运行后输出到 `q4_output/`，包含站点级方案和按服务标价表。

使用：直接运行 `python q4_solution.py`。
可选参数：--config 用于指定 JSON 配置文件（覆盖顶部常量）。
"""

BASE_DIR = Path(__file__).resolve().parent
Q2_DIR = BASE_DIR / "q2_output"
Q1_DIR = BASE_DIR / "q1_output"
OUTPUT_DIR = BASE_DIR / "q4_output"

# ========== 可修改参数（第四问变动处） ==========
# 搜索的统一 markup 候选集（例如允许更大折扣/更高提价）
MARKUP_OPTIONS = [-0.3, -0.2, -0.1, 0.0, 0.05, 0.10, 0.15, 0.20, 0.30]

# 每站每日补贴上限（可按题意调整）
CAP_PER_SIZE = {"小型": 1100.0, "中型": 2000.0, "大型": 3000.0}

# 利润率上限（百分比），如果解超出该值则视为不可行
PROFIT_RATE_LIMIT = 8.0

# 日固定成本近似（可用附件3 精确替换）
SIZE_TO_DAILY_FIXED = {"小型": 2000.0, "中型": 3200.0, "大型": 4400.0}
# ==================================================


def load_q2_results():
    station_df = pd.read_csv(Q2_DIR / "best_station_plan.csv", encoding="utf-8-sig")
    assign_df = pd.read_csv(Q2_DIR / "best_community_assignment.csv", encoding="utf-8-sig")
    return station_df, assign_df


def load_q1_theoretical():
    theo = pd.read_csv(Q1_DIR / "year5_theoretical_demand.csv", encoding="utf-8-sig")
    return theo


def load_service_baseline_and_costs():
    wb = load_workbook(BASE_DIR / "附件2：服务需求数据.xlsx", data_only=True)
    revenue = wb[wb.sheetnames[1]]
    prices = {}
    costs = {}
    for r in range(3, 9):
        svc = revenue.cell(r, 1).value
        val = revenue.cell(r, 2).value
        cost = revenue.cell(r, 3).value
        try:
            prices[svc] = float(val)
        except Exception:
            prices[svc] = 0.0
        try:
            costs[svc] = float(cost)
        except Exception:
            costs[svc] = 0.0
    return prices, costs


def distance_satisfaction(distance: float) -> float:
    if distance <= 300:
        return 1.0
    if distance <= 500:
        return 0.90
    if distance <= 650:
        return 0.75
    if distance <= 1000:
        return 0.60
    return 0.0


def tier_from_utilization(util: float) -> float:
    if util <= 0.60:
        return 1.00
    if util <= 0.75:
        return 0.93
    if util <= 0.85:
        return 0.85
    if util <= 0.95:
        return 0.72
    return 0.60


def price_satisfaction(baseline: float, net_price: float) -> float:
    if net_price <= baseline:
        return 1.00
    ratio = (net_price - baseline) / baseline
    if ratio <= 0.10:
        return 0.90
    if ratio <= 0.20:
        return 0.75
    return 0.60


def evaluate_markups(station_df, assign_df, theo, baseline_prices, service_costs):
    station_info = {}
    for _, row in station_df.iterrows():
        site = row["站点"]
        station_info[site] = {
            "规模": row["规模"],
            "日容量": float(row["日容量"]),
            "assigned": []
        }

    for _, row in assign_df.iterrows():
        station = row["分配站点"]
        if pd.isna(station):
            continue
        station_info[station]["assigned"].append(row["小区"]) 

    theo_total = theo.groupby(["小区"]) ["理论月需求次数"].sum().to_dict()
    theo_service = {}
    for _, r in theo.iterrows():
        theo_service.setdefault(r["小区"], {})[r["服务项目"]] = float(r["理论月需求次数"]) 

    sites = list(station_info.keys())
    best = None

    for markup_choice in product(MARKUP_OPTIONS, repeat=len(sites)):
        markups = {site: m for site, m in zip(sites, markup_choice)}
        station_s2 = {site: 0.6 for site in sites}

        # fixed-point iteration
        for _ in range(12):
            station_monthly_effective = {site: 0.0 for site in sites}
            station_monthly_non_emergency = {site: 0.0 for site in sites}

            for site in sites:
                for comm in station_info[site]["assigned"]:
                    # distance S1
                    dist = float(assign_df[assign_df["小区"] == comm]["距离(米)"].iloc[0])
                    s1 = distance_satisfaction(dist)
                    s2 = station_s2[site]
                    svc_dict = theo_service.get(comm, {})
                    if not svc_dict:
                        continue
                    total_units = sum(svc_dict.values())
                    if total_units <= 0:
                        continue
                    avg_baseline = sum(baseline_prices.get(svc, 0.0) * cnt for svc, cnt in svc_dict.items()) / total_units
                    avg_price = avg_baseline * (1 + markups[site])
                    s3 = price_satisfaction(avg_baseline, avg_price)
                    S = 0.2 * s1 + 0.3 * s2 + 0.5 * s3
                    station_monthly_effective[site] += sum(svc_dict.values()) * S
                    station_monthly_non_emergency[site] += sum(cnt for svc, cnt in svc_dict.items() if svc != "紧急救助") * S

            new_s2 = {}
            for site in sites:
                daily = station_info[site]["日容量"]
                util = station_monthly_effective[site] / (30.0 * daily) if daily > 0 else 0.0
                new_s2[site] = tier_from_utilization(util)
            if new_s2 == station_s2:
                break
            station_s2 = new_s2

        # compute profits and check feasibility
        station_annual_profit = {}
        station_satisfaction = {}
        feasible = True
        for site in sites:
            assigned = station_info[site]["assigned"]
            monthly_non_em = 0.0
            annual_revenue = 0.0
            annual_cost = 0.0
            community_satisfaction = {}
            for comm in assigned:
                dist = float(assign_df[assign_df["小区"] == comm]["距离(米)"].iloc[0])
                s1 = distance_satisfaction(dist)
                s2 = station_s2[site]
                svc_dict = theo_service.get(comm, {})
                comm_total = sum(svc_dict.values())
                if comm_total <= 0:
                    continue
                markup = markups[site]
                avg_baseline = sum(baseline_prices.get(svc, 0.0) * cnt for svc, cnt in svc_dict.items()) / comm_total
                avg_price = avg_baseline * (1 + markup)
                s3 = price_satisfaction(avg_baseline, avg_price)
                S = 0.2 * s1 + 0.3 * s2 + 0.5 * s3
                community_satisfaction[comm] = S
                monthly_non_em += sum(cnt for svc, cnt in svc_dict.items() if svc != "紧急救助") * S
                for svc, monthly_cnt in svc_dict.items():
                    eff_monthly = monthly_cnt * S
                    price_set = baseline_prices.get(svc, 0.0) * (1 + markup)
                    annual_revenue += eff_monthly * 12.0 * price_set
                    annual_cost += eff_monthly * 12.0 * service_costs.get(svc, 0.0)

            daily_non_em = monthly_non_em / 30.0 if monthly_non_em > 0 else 0.0
            cap = CAP_PER_SIZE.get(station_info[site]["规模"], 0.0)
            per_person_subsidy = min(2.0, cap / daily_non_em) if daily_non_em > 0 else 0.0
            annual_subsidy = per_person_subsidy * monthly_non_em / 30.0 * 365.0

            daily_fixed = SIZE_TO_DAILY_FIXED.get(station_info[site]["规模"], 2000.0)
            annual_operation_cost = daily_fixed * 365.0 + 10000.0 / 20.0 * (18.0 if station_info[site]["规模"] == "小型" else 32.0 if station_info[site]["规模"] == "中型" else 45.0)

            service_profit = annual_revenue - annual_cost
            profit_rate = (service_profit + annual_subsidy - annual_operation_cost) / annual_operation_cost * 100.0
            station_annual_profit[site] = service_profit + annual_subsidy - annual_operation_cost

            station_satisfaction[site] = (
                sum(community_satisfaction.get(c, 0.0) * theo_total.get(c, 0.0) for c in assigned)
                / sum(theo_total.get(c, 0.0) for c in assigned)
                if assigned
                else 0.0
            )

            if profit_rate > PROFIT_RATE_LIMIT:
                feasible = False
                break

        if not feasible:
            continue

        # overall metrics
        pop = q1_solution.load_community_data()[0]
        pop_map = {row["小区"]: float(row["60+"]) for _, row in pop.iterrows()}
        numerator = 0.0
        denom = 0.0
        for site in sites:
            for c in station_info[site]["assigned"]:
                numerator += station_satisfaction.get(site, 0.0) * pop_map.get(c, 0.0)
                denom += pop_map.get(c, 0.0)

        avg_satisfaction = numerator / denom if denom else 0.0
        total_profit = sum(station_annual_profit.values())

        candidate = {
            "markups": markups,
            "avg_satisfaction": avg_satisfaction,
            "total_profit": total_profit,
            "station_profit": station_annual_profit,
            "station_satisfaction": station_satisfaction,
        }

        if best is None or (candidate["avg_satisfaction"] > best["avg_satisfaction"] or (abs(candidate["avg_satisfaction"]-best["avg_satisfaction"])<1e-6 and candidate["total_profit"]>best["total_profit"])):
            best = candidate

    return best


def main():
    station_df, assign_df = load_q2_results()
    theo = load_q1_theoretical()
    baseline_prices, service_costs = load_service_baseline_and_costs()

    OUTPUT_DIR.mkdir(exist_ok=True)

    best = evaluate_markups(station_df, assign_df, theo, baseline_prices, service_costs)
    if best is None:
        print("未找到满足条件的方案（请调整参数）。")
        return

    out_rows = []
    for site, m in best["markups"].items():
        out_rows.append({"站点": site, "统一markup": m, "站点满意度": best["station_satisfaction"].get(site, 0.0), "年利润(元)": best["station_profit"].get(site, 0.0)})
    pd.DataFrame(out_rows).to_csv(OUTPUT_DIR / "best_pricing_plan.csv", index=False, encoding="utf-8-sig")

    # per-service pricing
    svc_rows = []
    wb = load_workbook(BASE_DIR / "附件2：服务需求数据.xlsx", data_only=True)
    sheet = wb[wb.sheetnames[1]]
    services = []
    for r in range(3, 9):
        svc = sheet.cell(r, 1).value
        price = sheet.cell(r, 2).value
        try:
            price = float(price)
        except Exception:
            price = 0.0
        services.append((svc, price))

    for site, m in best["markups"].items():
        for svc, base in services:
            svc_rows.append({"站点": site, "服务": svc, "基准价": base, "统一markup": m, "标价": round(base * (1 + m), 2)})

    pd.DataFrame(svc_rows).to_csv(OUTPUT_DIR / "best_pricing_per_service.csv", index=False, encoding="utf-8-sig")
    print("已保存 q4 输出：", OUTPUT_DIR)


if __name__ == "__main__":
    main()
