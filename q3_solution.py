from __future__ import annotations

from itertools import product
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
Q2_DIR = BASE_DIR / "q2_output"
Q1_DIR = BASE_DIR / "q1_output"
OUTPUT_DIR = BASE_DIR / "q3_output"


SERVICE_MARKUP_OPTIONS = [-0.2, -0.1, 0.0, 0.05, 0.10, 0.15, 0.25, 0.35, 0.50]
PAID_SERVICES = ["助餐", "日间照料", "上门护理", "康复理疗", "助浴"]
EMERGENCY_SERVICE = "紧急救助"

CAP_PER_SIZE = {"小型": 1000.0, "中型": 1800.0, "大型": 2600.0}
SIZE_TO_DAILY_FIXED = {"小型": 2000.0, "中型": 3200.0, "大型": 4400.0}
SIZE_TO_CONSTRUCTION_WAN = {"小型": 18.0, "中型": 32.0, "大型": 45.0}


def load_q2_results() -> tuple[pd.DataFrame, pd.DataFrame]:
    station_df = pd.read_csv(Q2_DIR / "best_station_plan.csv", encoding="utf-8-sig")
    assign_df = pd.read_csv(Q2_DIR / "best_community_assignment.csv", encoding="utf-8-sig")
    return station_df, assign_df


def load_year5_inputs() -> tuple[pd.DataFrame, pd.DataFrame]:
    population = pd.read_csv(Q1_DIR / "year5_population.csv", encoding="utf-8-sig")
    actual_demand = pd.read_csv(Q1_DIR / "year5_actual_demand.csv", encoding="utf-8-sig")
    return population, actual_demand


def load_service_baseline_and_costs() -> tuple[dict[str, float], dict[str, float]]:
    workbook = load_workbook(BASE_DIR / "附件2：服务需求数据.xlsx", data_only=True)
    revenue_sheet = workbook[workbook.sheetnames[1]]

    prices: dict[str, float] = {}
    costs: dict[str, float] = {}
    for row in range(3, 9):
        service = revenue_sheet.cell(row, 1).value
        price_value = revenue_sheet.cell(row, 2).value
        cost_value = revenue_sheet.cell(row, 3).value
        try:
            prices[service] = float(price_value)
        except Exception:
            prices[service] = 0.0
        try:
            costs[service] = float(cost_value)
        except Exception:
            costs[service] = 0.0
    return prices, costs


def distance_satisfaction(distance: float) -> float:
    if distance <= 300:
        return 1.0
    if distance <= 500:
        return 0.90
    if distance <= 650:
        return 0.75
    if distance <= 1000:
        return 0.60
    return 0.0


def tier_from_utilization(utilization: float) -> float:
    if utilization <= 0.60:
        return 1.00
    if utilization <= 0.75:
        return 0.93
    if utilization <= 0.85:
        return 0.85
    if utilization <= 0.95:
        return 0.72
    return 0.60


def price_satisfaction(baseline: float, net_price: float) -> float:
    if baseline <= 0:
        return 1.0
    if net_price <= baseline:
        return 1.0
    ratio = (net_price - baseline) / baseline
    if ratio <= 0.10:
        return 0.90
    if ratio <= 0.20:
        return 0.75
    return 0.60


def build_station_inputs(
    station_df: pd.DataFrame,
    assign_df: pd.DataFrame,
    population: pd.DataFrame,
    actual_demand: pd.DataFrame,
) -> tuple[dict[str, dict[str, object]], dict[str, float]]:
    population_map = {row["小区"]: float(row["60+"]) for _, row in population.iterrows()}

    community_service_demand: dict[str, dict[str, float]] = {}
    community_total_demand: dict[str, float] = {}
    community_non_emergency_demand: dict[str, float] = {}
    for _, row in actual_demand.iterrows():
        community = row["小区"]
        service = row["服务项目"]
        demand = float(row["实际月均需求次数"])
        community_service_demand.setdefault(community, {})[service] = demand
        community_total_demand[community] = community_total_demand.get(community, 0.0) + demand
        if service != EMERGENCY_SERVICE:
            community_non_emergency_demand[community] = community_non_emergency_demand.get(community, 0.0) + demand

    station_info: dict[str, dict[str, object]] = {}
    for _, row in station_df.iterrows():
        site = str(row["站点"])
        station_info[site] = {
            "规模": str(row["规模"]),
            "日容量": float(row["日容量"]),
            "assigned": [],
        }

    for _, row in assign_df.iterrows():
        station = row["分配站点"]
        if pd.isna(station):
            continue
        station_info[str(station)]["assigned"].append(str(row["小区"]))

    for site, info in station_info.items():
        info["community_service_demand"] = {
            community: community_service_demand.get(community, {})
            for community in info["assigned"]
        }
        info["community_total_demand"] = {
            community: community_total_demand.get(community, 0.0)
            for community in info["assigned"]
        }
        info["community_non_emergency_demand"] = {
            community: community_non_emergency_demand.get(community, 0.0)
            for community in info["assigned"]
        }
        info["population_weight"] = {community: population_map.get(community, 0.0) for community in info["assigned"]}

    return station_info, population_map


def evaluate_station_prices(
    site: str,
    station_info: dict[str, object],
    baseline_prices: dict[str, float],
    service_costs: dict[str, float],
) -> dict[str, object] | None:
    size_name = str(station_info["规模"])
    daily_capacity = float(station_info["日容量"])
    assigned: list[str] = list(station_info["assigned"])
    community_service_demand: dict[str, dict[str, float]] = station_info["community_service_demand"]
    community_total_demand: dict[str, float] = station_info["community_total_demand"]
    community_non_emergency_demand: dict[str, float] = station_info["community_non_emergency_demand"]
    population_weight: dict[str, float] = station_info["population_weight"]

    fixed_services = [EMERGENCY_SERVICE]
    paid_services = [service for service in PAID_SERVICES if service in baseline_prices]
    if not paid_services:
        return None

    best_result: dict[str, object] | None = None

    for markup_choice in product(SERVICE_MARKUP_OPTIONS, repeat=len(paid_services)):
        service_markups = {service: markup for service, markup in zip(paid_services, markup_choice)}
        service_prices = {
            service: baseline_prices[service] * (1.0 + service_markups.get(service, 0.0))
            for service in paid_services
        }
        service_prices[EMERGENCY_SERVICE] = 0.0

        station_s2 = 0.60
        per_person_subsidy = 2.0

        for _ in range(20):
            monthly_effective = 0.0
            monthly_non_emergency_effective = 0.0
            community_detail_rows: list[dict[str, object]] = []

            for community in assigned:
                service_dict = community_service_demand.get(community, {})
                total_demand = community_total_demand.get(community, 0.0)
                non_emergency_demand = community_non_emergency_demand.get(community, 0.0)
                if total_demand <= 0:
                    continue

                s1 = distance_satisfaction(float(station_info.get("distance_map", {}).get(community, 0.0)))

                price_score_sum = 0.0
                price_weight_sum = 0.0
                service_price_scores: dict[str, float] = {}
                for service in paid_services:
                    demand_count = float(service_dict.get(service, 0.0))
                    if demand_count <= 0:
                        continue
                    net_price = max(0.0, service_prices[service] - per_person_subsidy)
                    score = price_satisfaction(baseline_prices[service], net_price)
                    service_price_scores[service] = score
                    price_score_sum += demand_count * score
                    price_weight_sum += demand_count

                s3 = price_score_sum / price_weight_sum if price_weight_sum > 0 else 1.0
                community_satisfaction = 0.2 * s1 + 0.3 * station_s2 + 0.5 * s3

                monthly_effective += total_demand * community_satisfaction
                monthly_non_emergency_effective += non_emergency_demand * community_satisfaction

                community_detail_rows.append(
                    {
                        "小区": community,
                        "距离满意度S1": round(s1, 4),
                        "响应满意度S2": round(station_s2, 4),
                        "价格满意度S3": round(s3, 4),
                        "社区满意度": round(community_satisfaction, 4),
                        "理论月需求次数": round(total_demand, 2),
                        "非紧急月需求次数": round(non_emergency_demand, 2),
                    }
                )

            utilization = monthly_effective / (30.0 * daily_capacity) if daily_capacity > 0 else 0.0
            next_s2 = tier_from_utilization(utilization)

            daily_non_emergency = monthly_non_emergency_effective / 30.0
            cap = CAP_PER_SIZE.get(size_name, 0.0)
            next_subsidy = min(2.0, cap / daily_non_emergency) if daily_non_emergency > 0 else 0.0

            if abs(next_subsidy - per_person_subsidy) < 1e-9 and next_s2 == station_s2:
                station_s2 = next_s2
                per_person_subsidy = next_subsidy
                break

            station_s2 = next_s2
            per_person_subsidy = next_subsidy

        station_annual_revenue = 0.0
        station_annual_direct_cost = 0.0
        station_annual_subsidy = 0.0
        station_community_rows: list[dict[str, object]] = []
        station_population_weighted_sum = 0.0
        station_population_weight = 0.0

        for community in assigned:
            service_dict = community_service_demand.get(community, {})
            total_demand = community_total_demand.get(community, 0.0)
            non_emergency_demand = community_non_emergency_demand.get(community, 0.0)
            if total_demand <= 0:
                continue

            s1 = distance_satisfaction(float(station_info.get("distance_map", {}).get(community, 0.0)))
            price_score_sum = 0.0
            price_weight_sum = 0.0
            service_price_scores: dict[str, float] = {}

            for service in paid_services:
                demand_count = float(service_dict.get(service, 0.0))
                if demand_count <= 0:
                    continue
                net_price = max(0.0, service_prices[service] - per_person_subsidy)
                score = price_satisfaction(baseline_prices[service], net_price)
                service_price_scores[service] = score
                price_score_sum += demand_count * score
                price_weight_sum += demand_count

            s3 = price_score_sum / price_weight_sum if price_weight_sum > 0 else 1.0
            community_satisfaction = 0.2 * s1 + 0.3 * station_s2 + 0.5 * s3

            station_population_weighted_sum += community_satisfaction * population_weight.get(community, 0.0)
            station_population_weight += population_weight.get(community, 0.0)

            station_community_rows.append(
                {
                    "站点": site,
                    "小区": community,
                    "距离(米)": float(station_info.get("distance_map", {}).get(community, 0.0)),
                    "距离满意度S1": round(s1, 4),
                    "响应满意度S2": round(station_s2, 4),
                    "价格满意度S3": round(s3, 4),
                    "社区满意度": round(community_satisfaction, 4),
                }
            )

            for service, monthly_count in service_dict.items():
                effective_monthly = float(monthly_count) * community_satisfaction
                price = float(service_prices.get(service, 0.0))
                baseline = float(baseline_prices.get(service, 0.0))
                annual_revenue = effective_monthly * 12.0 * price
                annual_cost = effective_monthly * 12.0 * float(service_costs.get(service, 0.0))
                station_annual_revenue += annual_revenue
                station_annual_direct_cost += annual_cost

                if service != EMERGENCY_SERVICE:
                    station_annual_subsidy += per_person_subsidy * effective_monthly * 12.0

        annual_operation_cost = (
            SIZE_TO_DAILY_FIXED.get(size_name, 2000.0) * 365.0
            + SIZE_TO_CONSTRUCTION_WAN.get(size_name, 18.0) * 10000.0 / 20.0
        )
        service_profit = station_annual_revenue - station_annual_direct_cost
        annual_profit = service_profit + station_annual_subsidy - annual_operation_cost
        profit_rate = (annual_profit / annual_operation_cost * 100.0) if annual_operation_cost > 0 else float("inf")
        avg_satisfaction = (
            station_population_weighted_sum / station_population_weight if station_population_weight > 0 else 0.0
        )

        if profit_rate < 0.0 or profit_rate > 8.0:
            continue

        candidate = {
            "站点": site,
            "规模": size_name,
            "markup": service_markups,
            "service_prices": service_prices.copy(),
            "response_s2": station_s2,
            "avg_satisfaction": avg_satisfaction,
            "station_profit": annual_profit,
            "service_profit": service_profit,
            "annual_subsidy": station_annual_subsidy,
            "annual_operation_cost": annual_operation_cost,
            "profit_rate": profit_rate,
            "community_rows": station_community_rows,
            "service_rows": [
                {
                    "站点": site,
                    "服务": service,
                    "基准价": round(float(baseline_prices.get(service, 0.0)), 2),
                    "最优定价": round(float(service_prices.get(service, 0.0)), 2),
                    "markup": round(float(service_markups.get(service, 0.0)), 4),
                }
                for service in paid_services + [EMERGENCY_SERVICE]
            ],
        }

        if best_result is None or (
            candidate["avg_satisfaction"] > best_result["avg_satisfaction"]
            or (
                abs(candidate["avg_satisfaction"] - best_result["avg_satisfaction"]) < 1e-9
                and candidate["station_profit"] > best_result["station_profit"]
            )
        ):
            best_result = candidate

    return best_result


def attach_station_distances(station_info: dict[str, dict[str, object]], assign_df: pd.DataFrame) -> None:
    for site, info in station_info.items():
        distance_map: dict[str, float] = {}
        for community in info["assigned"]:
            row = assign_df[assign_df["小区"] == community]
            if row.empty:
                distance_map[community] = 0.0
            else:
                distance_map[community] = float(row["距离(米)"].iloc[0])
        info["distance_map"] = distance_map


def main() -> None:
    station_df, assign_df = load_q2_results()
    population_df, actual_demand_df = load_year5_inputs()
    baseline_prices, service_costs = load_service_baseline_and_costs()

    station_info, _ = build_station_inputs(station_df, assign_df, population_df, actual_demand_df)
    attach_station_distances(station_info, assign_df)

    OUTPUT_DIR.mkdir(exist_ok=True)

    all_station_rows: list[dict[str, object]] = []
    all_service_rows: list[dict[str, object]] = []
    all_community_rows: list[dict[str, object]] = []

    for site, info in station_info.items():
        result = evaluate_station_prices(site, info, baseline_prices, service_costs)
        if result is None:
            raise RuntimeError(f"站点 {site} 未找到满足利润率约束的定价方案。")

        all_station_rows.append(
            {
                "站点": site,
                "规模": result["规模"],
                "助餐": result["service_prices"].get("助餐", 0.0),
                "日间照料": result["service_prices"].get("日间照料", 0.0),
                "上门护理": result["service_prices"].get("上门护理", 0.0),
                "康复理疗": result["service_prices"].get("康复理疗", 0.0),
                "助浴": result["service_prices"].get("助浴", 0.0),
                "紧急救助": result["service_prices"].get("紧急救助", 0.0),
                "响应满意度S2": round(result["response_s2"], 4),
                "平均满意度": round(result["avg_satisfaction"], 4),
                "政府补贴(元)": round(result["annual_subsidy"], 2),
                "年运营成本(元)": round(result["annual_operation_cost"], 2),
                "年度利润(元)": round(result["station_profit"], 2),
                "利润率(%)": round(result["profit_rate"], 4),
            }
        )

        all_service_rows.extend(result["service_rows"])
        all_community_rows.extend(result["community_rows"])

    station_output = pd.DataFrame(all_station_rows)
    service_output = pd.DataFrame(all_service_rows)
    community_output = pd.DataFrame(all_community_rows)

    station_output.to_csv(OUTPUT_DIR / "best_pricing_plan.csv", index=False, encoding="utf-8-sig")
    service_output.to_csv(OUTPUT_DIR / "best_pricing_per_service.csv", index=False, encoding="utf-8-sig")
    community_output.to_csv(OUTPUT_DIR / "best_community_satisfaction.csv", index=False, encoding="utf-8-sig")

    print("最优定价方案已生成：")
    print(station_output.to_string(index=False))


if __name__ == "__main__":
    main()