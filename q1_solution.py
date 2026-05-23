from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# 第 1 问说明：
# 1. 附件 1 的第一个工作表给出各小区当前人口结构。
# 2. 附件 1 的第二个工作表给出转移概率，其中 B3=0.045 表示“自理 -> 半失能”，B4=0.10 表示“半失能 -> 失能”。
# 3. 附件 2 给出各类老人对不同服务的月均需求次数、服务基准价格和月消费上限。
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "q1_output"


def load_community_data() -> tuple[pd.DataFrame, float, float]:
    workbook = load_workbook(BASE_DIR / "附件1：小区基础数据.xlsx", data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    communities = pd.DataFrame(
        list(worksheet.iter_rows(min_row=3, max_row=12, values_only=True)),
        columns=["小区", "总人口", "60+", "自理", "半失能", "失能", "人均月收入"],
    )

    transition_sheet = workbook[workbook.sheetnames[1]]
    # 这里直接读取题目给出的转移概率：
    # 自理老人每年有 4.5% 转移为半失能，半失能老人每年有 10% 转移为失能。
    self_to_semi = 0.045
    semi_to_disabled = 0.1
    return communities, self_to_semi, semi_to_disabled


def load_service_data() -> tuple[pd.DataFrame, dict[str, float], dict[str, float]]:
    workbook = load_workbook(BASE_DIR / "附件2：服务需求数据.xlsx", data_only=True)

    demand_sheet = workbook[workbook.sheetnames[0]]
    service_names = [demand_sheet.cell(row, 1).value for row in range(3, 9)]
    demand_rates = pd.DataFrame(
        {
            "自理": [demand_sheet.cell(row, 2).value for row in range(3, 9)],
            "半失能": [demand_sheet.cell(row, 3).value for row in range(3, 9)],
            "失能": [demand_sheet.cell(row, 4).value for row in range(3, 9)],
        },
        index=service_names,
    )

    revenue_sheet = workbook[workbook.sheetnames[1]]
    service_prices = {
        revenue_sheet.cell(row, 1).value: (0.0 if row == 8 else float(revenue_sheet.cell(row, 2).value))
        for row in range(3, 9)
    }

    cap_ratios = {
        "自理": 0.20,
        "半失能": 0.25,
        "失能": 0.30,
    }
    return demand_rates, service_prices, cap_ratios


def forecast_five_years(
    communities: pd.DataFrame,
    self_to_semi: float,
    semi_to_disabled: float,
) -> list[pd.DataFrame]:
    result = communities.copy()
    yearly_results: list[pd.DataFrame] = []
    for _ in range(5):
        # 每年先扣除自然死亡，再叠加 60+ 新增人口；
        # 新增人口按题意计入自理老人。
        new_60_plus = 0.07 * result["60+"]
        next_self = 0.95 * result["自理"] * (1 - self_to_semi) + new_60_plus
        next_semi = 0.95 * (
            result["自理"] * self_to_semi + result["半失能"] * (1 - semi_to_disabled)
        )
        next_disabled = 0.95 * (result["半失能"] * semi_to_disabled + result["失能"])

        result[["自理", "半失能", "失能"]] = pd.DataFrame(
            {"自理": next_self, "半失能": next_semi, "失能": next_disabled}
        )
        result["60+"] = result[["自理", "半失能", "失能"]].sum(axis=1)
        yearly_results.append(result.copy())

    return yearly_results


def build_service_tables(
    forecast: pd.DataFrame,
    communities: pd.DataFrame,
    demand_rates: pd.DataFrame,
    service_prices: dict[str, float],
    cap_ratios: dict[str, float],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    theoretical_rows: list[dict[str, object]] = []
    actual_rows: list[dict[str, object]] = []

    for _, community_row in forecast.iterrows():
        community = community_row["小区"]
        income = float(communities.loc[communities["小区"] == community, "人均月收入"].iloc[0])

        for elder_type, cap_key in [("自理", "自理"), ("半失能", "半失能"), ("失能", "失能")]:
            elder_count = float(community_row[elder_type])
            # 理论需求 = 老人数量 × 对应服务的月均需求次数。
            theoretical_demands = {
                service: elder_count * float(demand_rates.loc[service, elder_type])
                for service in demand_rates.index
            }

            # 若理论费用超过消费上限，则按照题意对各项服务次数等比例缩减。
            theoretical_fee = sum(
                theoretical_demands[service] * float(service_prices[service])
                for service in demand_rates.index
            )
            monthly_cap = income * float(cap_ratios[cap_key]) * elder_count
            scale = 1.0 if theoretical_fee <= monthly_cap else monthly_cap / theoretical_fee

            for service in demand_rates.index:
                theoretical_rows.append(
                    {
                        "小区": community,
                        "老人类型": elder_type,
                        "服务项目": service,
                        "理论月需求次数": round(theoretical_demands[service], 2),
                    }
                )
                actual_rows.append(
                    {
                        "小区": community,
                        "老人类型": elder_type,
                        "服务项目": service,
                        "消费约束缩放系数": round(scale, 4),
                        "实际月均需求次数": round(theoretical_demands[service] * scale),
                    }
                )

    return pd.DataFrame(theoretical_rows), pd.DataFrame(actual_rows)


def main() -> None:
    communities, self_to_semi, semi_to_disabled = load_community_data()
    demand_rates, service_prices, cap_ratios = load_service_data()

    yearly_forecasts = forecast_five_years(communities, self_to_semi, semi_to_disabled)
    forecast = yearly_forecasts[-1]
    theoretical, actual = build_service_tables(
        forecast=forecast,
        communities=communities,
        demand_rates=demand_rates,
        service_prices=service_prices,
        cap_ratios=cap_ratios,
    )

    OUTPUT_DIR.mkdir(exist_ok=True)
    forecast_out = forecast[["小区", "自理", "半失能", "失能", "60+"]].copy()
    forecast_out[["自理", "半失能", "失能", "60+"]] = forecast_out[["自理", "半失能", "失能", "60+"]].round(2)
    forecast_out.to_csv(OUTPUT_DIR / "year5_population.csv", index=False, encoding="utf-8-sig")
    for year, year_forecast in enumerate(yearly_forecasts, start=1):
        year_out = year_forecast[["小区", "自理", "半失能", "失能", "60+"]].copy()
        year_out[["自理", "半失能", "失能", "60+"]] = year_out[["自理", "半失能", "失能", "60+"]].round(2)
        year_out.to_csv(OUTPUT_DIR / f"year{year}_population.csv", index=False, encoding="utf-8-sig")
    theoretical.to_csv(OUTPUT_DIR / "year5_theoretical_demand.csv", index=False, encoding="utf-8-sig")
    actual.to_csv(OUTPUT_DIR / "year5_actual_demand.csv", index=False, encoding="utf-8-sig")

    for year, year_forecast in enumerate(yearly_forecasts, start=1):
        year_out = year_forecast[["小区", "自理", "半失能", "失能", "60+"]].copy()
        year_out[["自理", "半失能", "失能", "60+"]] = year_out[["自理", "半失能", "失能", "60+"]].round(2)
        print(f"第{year}年末老人数量预测：")
        print(year_out.to_string(index=False))
        print()

    print("第5年末理论月需求（前20行）：")
    print(theoretical.head(20).to_string(index=False))
    print()
    print("第5年末消费约束后的实际月均需求（前20行）：")
    print(actual.head(20).to_string(index=False))


if __name__ == "__main__":
    main()