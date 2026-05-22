from __future__ import annotations

from dataclasses import dataclass
from itertools import combinations, product
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import q1_solution


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "q2_output"


SIZE_OPTIONS = [
    ("小型", 18.0, 2000.0, 1000),
    ("中型", 32.0, 3200.0, 2000),
    ("大型", 45.0, 4400.0, 3000),
]

S2_LEVELS = [1.00, 0.93, 0.85, 0.72, 0.60]
MAX_SELECTED_SITES = 5
TOP_CANDIDATE_SITES = 6


@dataclass(frozen=True)
class StationPlan:
    site: str
    size_name: str
    construction_cost_wan: float
    annual_fixed_cost_yuan: float
    daily_capacity: float


def load_distance_matrix() -> pd.DataFrame:
    """读取附件 4 的小区间距离矩阵。"""
    workbook = load_workbook(BASE_DIR / "附件4：小区间距离矩阵.xlsx", data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    names = [worksheet.cell(2, col).value for col in range(2, 12)]
    data = []
    for row in range(3, 13):
        data.append([worksheet.cell(row, col).value for col in range(2, 12)])
    return pd.DataFrame(data, index=names, columns=names, dtype=float)


def load_year5_demand() -> tuple[pd.DataFrame, pd.DataFrame]:
    """基于第 1 问的结果，整理第 5 年末的社区总需求和服务需求。"""
    communities, self_to_semi, semi_to_disabled = q1_solution.load_community_data()
    demand_rates, service_prices, cap_ratios = q1_solution.load_service_data()
    forecast = q1_solution.forecast_five_years(communities, self_to_semi, semi_to_disabled)
    _, actual = q1_solution.build_service_tables(
        forecast=forecast,
        communities=communities,
        demand_rates=demand_rates,
        service_prices=service_prices,
        cap_ratios=cap_ratios,
    )

    community_population = forecast[["小区", "60+"]].copy()
    community_population.rename(columns={"60+": "60岁以上人口"}, inplace=True)

    demand_total = (
        actual.groupby(["小区", "服务项目"], as_index=False)["实际月均需求次数"].sum()
    )
    return community_population, demand_total


def tier_from_utilization(utilization: float) -> float:
    """把利用率映射到题目规定的响应满意度 S2。"""
    if utilization <= 0.60:
        return 1.00
    if utilization <= 0.75:
        return 0.93
    if utilization <= 0.85:
        return 0.85
    if utilization <= 0.95:
        return 0.72
    return 0.60


def distance_satisfaction(distance: float) -> float:
    """把距离映射到题目规定的距离满意度 S1。"""
    if distance <= 300:
        return 1.00
    if distance <= 500:
        return 0.90
    if distance <= 650:
        return 0.75
    if distance <= 1000:
        return 0.60
    return 0.0


def build_station_plan(site: str, size_name: str) -> StationPlan:
    construction_cost, daily_fixed_cost, daily_capacity = next(
        (construction, fixed_cost, capacity)
        for name, construction, fixed_cost, capacity in SIZE_OPTIONS
        if name == size_name
    )
    return StationPlan(
        site=site,
        size_name=size_name,
        construction_cost_wan=construction_cost,
        annual_fixed_cost_yuan=daily_fixed_cost * 365,
        daily_capacity=daily_capacity,
    )


def feasible_size_names(required_capacity: float) -> list[str]:
    """根据理论日均需求，返回可覆盖该负载的站点规模。"""
    return [name for name, _, _, capacity in SIZE_OPTIONS if capacity >= required_capacity]


def solve_station_s2(
    assigned_communities: list[str],
    station_capacity: float,
    community_to_s1: dict[str, float],
    community_monthly_total: dict[str, float],
) -> tuple[float, float, float] | None:
    """在固定社区分配和容量下，求解站点响应满意度 S2 的固定点。"""
    if station_capacity <= 0:
        return None

    consistent_levels: list[tuple[float, float, float]] = []
    for s2 in S2_LEVELS:
        monthly_effective_load = 0.0
        for community in assigned_communities:
            s1 = community_to_s1[community]
            community_satisfaction = 0.2 * s1 + 0.3 * s2 + 0.5
            monthly_effective_load += community_monthly_total[community] * community_satisfaction

        utilization = monthly_effective_load / (30.0 * station_capacity)
        if tier_from_utilization(utilization) == s2:
            consistent_levels.append((s2, utilization, monthly_effective_load))

    if consistent_levels:
        return max(consistent_levels, key=lambda item: item[0])

    # 若没有直接的稳定解，则从最高满意度开始迭代，直到收敛。
    current_s2 = 1.00
    for _ in range(20):
        monthly_effective_load = 0.0
        for community in assigned_communities:
            s1 = community_to_s1[community]
            community_satisfaction = 0.2 * s1 + 0.3 * current_s2 + 0.5
            monthly_effective_load += community_monthly_total[community] * community_satisfaction

        utilization = monthly_effective_load / (30.0 * station_capacity)
        next_s2 = tier_from_utilization(utilization)
        if next_s2 == current_s2:
            return current_s2, utilization, monthly_effective_load
        current_s2 = next_s2

    return None


def evaluate_plan(
    selected_sites: list[str],
    size_names: dict[str, str],
    distance_matrix: pd.DataFrame,
    community_population: pd.DataFrame,
    demand_total: pd.DataFrame,
    service_prices: dict[str, float],
    service_costs: dict[str, float],
) -> dict[str, object] | None:
    """给定站点集合和规模方案，计算覆盖率、满意度、利润等指标。"""
    community_list = list(distance_matrix.index)
    community_60_plus = {
        row["小区"]: float(row["60岁以上人口"])
        for _, row in community_population.iterrows()
    }
    community_monthly_total = (
        demand_total.groupby("小区")["实际月均需求次数"].sum().to_dict()
    )
    community_service_demand = {
        community: demand_total[demand_total["小区"] == community].set_index("服务项目")["实际月均需求次数"].to_dict()
        for community in community_list
    }

    # 初始分配：按最近的可达站点分配。
    assignment: dict[str, str | None] = {}
    for community in community_list:
        reachable = [site for site in selected_sites if distance_matrix.loc[community, site] <= 1000]
        if not reachable:
            assignment[community] = None
            continue
        assignment[community] = min(reachable, key=lambda site: (distance_matrix.loc[community, site], site))

    station_plans = {site: build_station_plan(site, size_names[site]) for site in selected_sites}

    # 站点和社区之间存在“利用率 -> 响应满意度 -> 实际需求 -> 利用率”的闭环，用迭代法求稳定分配。
    for _ in range(12):
        station_s2: dict[str, float] = {}
        station_utilization: dict[str, float] = {}
        station_effective_monthly: dict[str, float] = {}
        station_annual_profit: dict[str, float] = {}
        station_community_satisfaction: dict[str, float] = {}

        valid = True
        for site in selected_sites:
            assigned = [community for community, assigned_site in assignment.items() if assigned_site == site]
            station_capacity = station_plans[site].daily_capacity
            community_to_s1 = {community: distance_satisfaction(distance_matrix.loc[community, site]) for community in assigned}
            result = solve_station_s2(assigned, station_capacity, community_to_s1, community_monthly_total)
            if result is None:
                valid = False
                break
            s2, utilization, monthly_effective_load = result
            station_s2[site] = s2
            station_utilization[site] = utilization
            station_effective_monthly[site] = monthly_effective_load

            annual_revenue = 0.0
            annual_direct_cost = 0.0
            for community in assigned:
                s1 = community_to_s1[community]
                community_satisfaction = 0.2 * s1 + 0.3 * s2 + 0.5
                station_community_satisfaction[community] = community_satisfaction
                for service, monthly_count in community_service_demand[community].items():
                    effective_yearly = float(monthly_count) * community_satisfaction * 12.0
                    annual_revenue += effective_yearly * float(service_prices[service])
                    annual_direct_cost += effective_yearly * float(service_costs[service])

            annual_fixed_cost = station_plans[site].annual_fixed_cost_yuan + station_plans[site].construction_cost_wan * 10000.0 / 20.0
            station_annual_profit[site] = annual_revenue - annual_direct_cost - annual_fixed_cost

        if not valid:
            return None

        new_assignment: dict[str, str | None] = {}
        for community in community_list:
            reachable = [site for site in selected_sites if distance_matrix.loc[community, site] <= 1000]
            if not reachable:
                new_assignment[community] = None
                continue

            # 社区选择自身满意度最高的站点，若满意度相同则选距离更近的站点。
            best_site = max(
                reachable,
                key=lambda site: (
                    0.2 * distance_satisfaction(distance_matrix.loc[community, site])
                    + 0.3 * station_s2[site]
                    + 0.5,
                    -distance_matrix.loc[community, site],
                    site,
                ),
            )
            new_assignment[community] = best_site

        if new_assignment == assignment:
            covered_population = sum(community_60_plus[community] for community, site in assignment.items() if site is not None)
            total_population = sum(community_60_plus.values())
            coverage = covered_population / total_population if total_population else 0.0

            covered_satisfactions = [
                0.2 * distance_satisfaction(distance_matrix.loc[community, assignment[community]])
                + 0.3 * station_s2[assignment[community]]
                + 0.5
                for community in community_list
                if assignment[community] is not None
            ]
            covered_weights = [community_60_plus[community] for community in community_list if assignment[community] is not None]
            avg_satisfaction = (
                sum(s * w for s, w in zip(covered_satisfactions, covered_weights)) / sum(covered_weights)
                if covered_weights
                else 0.0
            )
            total_profit = sum(station_annual_profit.values())

            community_result_rows = []
            for community in community_list:
                assigned_site = assignment[community]
                if assigned_site is None:
                    community_result_rows.append(
                        {
                            "小区": community,
                            "分配站点": None,
                            "距离(米)": None,
                            "距离满意度S1": 0.0,
                            "响应满意度S2": 0.0,
                            "社区满意度": 0.0,
                        }
                    )
                else:
                    community_result_rows.append(
                        {
                            "小区": community,
                            "分配站点": assigned_site,
                            "距离(米)": float(distance_matrix.loc[community, assigned_site]),
                            "距离满意度S1": distance_satisfaction(distance_matrix.loc[community, assigned_site]),
                            "响应满意度S2": station_s2[assigned_site],
                            "社区满意度": 0.2 * distance_satisfaction(distance_matrix.loc[community, assigned_site]) + 0.3 * station_s2[assigned_site] + 0.5,
                        }
                    )

            station_result_rows = []
            for site in selected_sites:
                assigned = [community for community, assigned_site in assignment.items() if assigned_site == site]
                station_result_rows.append(
                    {
                        "站点": site,
                        "规模": station_plans[site].size_name,
                        "分配社区": "、".join(assigned),
                        "覆盖社区数": len(assigned),
                        "日容量": station_plans[site].daily_capacity,
                        "利用率": round(station_utilization[site], 4),
                        "响应满意度S2": station_s2[site],
                        "年度利润(元)": round(station_annual_profit[site], 2),
                    }
                )

            return {
                "coverage": coverage,
                "avg_satisfaction": avg_satisfaction,
                "total_profit": total_profit,
                "assignment": assignment,
                "station_s2": station_s2,
                "station_utilization": station_utilization,
                "station_profit": station_annual_profit,
                "station_rows": pd.DataFrame(station_result_rows),
                "community_rows": pd.DataFrame(community_result_rows),
                "selected_sites": selected_sites,
                "size_names": size_names,
                "station_plans": station_plans,
            }

        assignment = new_assignment

    return None


def search_best_plan(
    distance_matrix: pd.DataFrame,
    community_population: pd.DataFrame,
    demand_total: pd.DataFrame,
    service_prices: dict[str, float],
    service_costs: dict[str, float],
) -> dict[str, object]:
    """穷举所有站点组合和规模组合，并按“覆盖率 -> 满意度 -> 利润”排序。"""
    sites = list(distance_matrix.index)
    ranked_sites = sorted(
        sites,
        key=lambda site: (
            sum(
                float(community_population.loc[community_population["小区"] == community, "60岁以上人口"].iloc[0])
                for community in sites
                if distance_matrix.loc[community, site] <= 1000
            ),
            -sites.index(site),
        ),
        reverse=True,
    )
    candidate_sites = ranked_sites[:TOP_CANDIDATE_SITES]
    print("用于穷举的候选站点：", candidate_sites)
    best_result: dict[str, object] | None = None

    for station_count in range(1, min(MAX_SELECTED_SITES, len(candidate_sites)) + 1):
        for selected_sites in combinations(candidate_sites, station_count):
            # 先做一个很便宜的下界剪枝：若所有站点都选最小规模，建设成本仍超预算，则直接跳过。
            if 18.0 * station_count > 120.0:
                continue

            for size_choice in product([option[0] for option in SIZE_OPTIONS], repeat=station_count):
                size_names = {site: size_name for site, size_name in zip(selected_sites, size_choice)}
                construction_cost = sum(
                    next(option[1] for option in SIZE_OPTIONS if option[0] == size_name)
                    for size_name in size_choice
                )
                if construction_cost > 120.0:
                    continue

                result = evaluate_plan(
                    selected_sites=list(selected_sites),
                    size_names=size_names,
                    distance_matrix=distance_matrix,
                    community_population=community_population,
                    demand_total=demand_total,
                    service_prices=service_prices,
                    service_costs=service_costs,
                )
                if result is None:
                    continue

                if best_result is None:
                    best_result = result
                    continue

                current_key = (result["coverage"], result["avg_satisfaction"], result["total_profit"])
                best_key = (best_result["coverage"], best_result["avg_satisfaction"], best_result["total_profit"])
                if current_key > best_key:
                    best_result = result

    if best_result is None:
        raise RuntimeError("没有找到满足预算约束的可行方案。")

    return best_result


def main() -> None:
    communities, _, _ = q1_solution.load_community_data()
    demand_rates, service_prices, _ = q1_solution.load_service_data()
    service_costs = {
        "助餐": 8.0,
        "日间照料": 16.0,
        "上门护理": 24.0,
        "康复理疗": 23.0,
        "助浴": 20.0,
        "紧急救助": 8.0,
    }

    distance_matrix = load_distance_matrix()
    community_population, demand_total = load_year5_demand()

    best = search_best_plan(
        distance_matrix=distance_matrix,
        community_population=community_population,
        demand_total=demand_total,
        service_prices=service_prices,
        service_costs=service_costs,
    )

    OUTPUT_DIR.mkdir(exist_ok=True)
    best["station_rows"].to_csv(OUTPUT_DIR / "best_station_plan.csv", index=False, encoding="utf-8-sig")
    best["community_rows"].to_csv(OUTPUT_DIR / "best_community_assignment.csv", index=False, encoding="utf-8-sig")

    print("最优方案：")
    print(f"站点数量：{len(best['selected_sites'])}")
    print(f"覆盖率：{best['coverage']:.4f}")
    print(f"平均满意度：{best['avg_satisfaction']:.4f}")
    print(f"总年度利润：{best['total_profit']:.2f} 元")
    print()
    print("站点明细：")
    print(best["station_rows"].to_string(index=False))
    print()
    print("社区分配明细：")
    print(best["community_rows"].to_string(index=False))


if __name__ == "__main__":
    main()